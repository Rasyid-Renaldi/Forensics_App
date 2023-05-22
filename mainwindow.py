from PySide6.QtCore import Qt, QStringListModel
from PySide6 import QtCore, QtGui, QtWidgets
from PySide6.QtCore import (QCoreApplication, QPropertyAnimation, QDate, QDateTime,
                            QMetaObject, QObject, QPoint, QRect, QSize, QTime, QUrl, Qt, QEvent)
from PySide6.QtGui import (QBrush, QColor, QConicalGradient, QCursor, QFont, QFontDatabase,
                           QIcon, QKeySequence, QLinearGradient, QPalette, QPainter, QPixmap, QRadialGradient)
from PySide6.QtWidgets import QApplication, QMainWindow, QMessageBox, QFileDialog, QLabel, QGraphicsDropShadowEffect, QProgressBar, QComboBox, QListView
import time
import enum
import os
import pandas as pd
from pytesseract import pytesseract
from PIL import Image
from openpyxl import Workbook

from main_ui import Ui_Aplikasi_Forensik
from splash_screen_ui import Ui_SplashScreen


class OS(enum.Enum):
    Windows = 1


class Language(enum.Enum):
    ENG = 'eng'
    IND = 'ind'


class Aplikasi_Forensik (QMainWindow, Ui_Aplikasi_Forensik):
    def __init__(self, app):
        super().__init__()
        self.setupUi(self)
        self.app = app
        self.text_file_path = None

        self.detectDevice.currentIndexChanged.connect(self.dialog_connect)
        # self.convertButton.clicked.connect(self.convert_image_to_excel)
        self.convertButton.clicked.connect(self.start_convert_image)

        self.actionQuit.triggered.connect(self.quit)
        self.actionAbout_App.triggered.connect(self.about)
        self.actionAbout_QT.triggered.connect(self.aboutQt)

        self.searchFile.clicked.connect(self.search_button)
        self.browse.clicked.connect(self.save_locations)

        self.scanButton.clicked.connect(self.dialog_scan)
        self.cekButton.clicked.connect(self.cek_button)

        self.activity_model = QStringListModel()
        self.viewData.setModel(self.activity_model)

        self.okButton.clicked.connect(self.exit)

    def start_convert_image(self):
        self.convertButton.setEnabled(False)

        self.add_activity("Mengubah gambar menjadi teks")
        self.add_activity("Path gambar: {}".format(self.search.text()))
        self.add_activity("Path penyimpanan: {}".format(self.text_file_path))
        self.add_activity("Nama examiner: {}".format(self.lineEdit.text()))
        self.add_activity("Lokasi file output: {}".format(self.save_locations()))

        animation = QPropertyAnimation(self.acquisitionBar, b"value")
        animation.setDuration(3000)
        animation.setStartValue(0)
        animation.setEndValue(100)
        # Menghubungkan ke fungsi convert_image_to_excel
        animation.finished.connect(self.convert_image_to_text)
        animation.start()

        self.convert_image_to_text()

    def convert_image_to_text(self):
        image_path = self.search.text()

        if image_path and self.text_file_path:
            image = Image.open(image_path)
            win_path = r'E:\Program Files\Tesseract-OCR\tesseract.exe'
            pytesseract.tesseract_cmd = win_path
            print('Run on: Windows\n')

            extracted_text = pytesseract.image_to_string(image)

            with open(self.text_file_path, "w", encoding="utf-8") as text_file:
                text_file.write(extracted_text)
            QMessageBox.information(
                self, "Ekstraksi Selesai", "Teks berhasil diekstrak dan disimpan dalam file teks"
            )
        else:
            QMessageBox.warning(
                self, "Error", "Mohon pilih file gambar dan lokasi penyimpanan terlebih dahulu!")

        self.convertButton.setEnabled(True)

    def quit(self):
        self.app.quit()

    def about(self):
        QMessageBox.information(
            self, "Aplikasi Forensik", "Aplikasi ini hanya bisa digunakan untuk kepentingan akuisisi data!")

    def aboutQt(self):
        QApplication.aboutQt()

    def search_button(self):
        search = QFileDialog.getOpenFileName(
            self, "Open File", "D:\Skripsi", "JPG files (*.jpg)")
        self.search.setText(search[0])
        self.add_activity("Path gambar: {}".format(search[0]))

    def save_locations(self):
        selected_location = QFileDialog.getSaveFileName(
            self, "Save File", "D:\Skripsi", "TXT files (*.txt)")
        self.browseLocation.setText(selected_location[0])
        self.text_file_path = selected_location[0]
        self.add_activity("Path penyimpanan: {}".format(selected_location[0]))

    def add_activity(self, activity):
        act = self.activity_model.stringList()
        act.append(activity)
        self.activity_model.setStringList(act)

    def dialog_scan(self):
        scan = QMessageBox.question(
            self, "Scan Button", "Ingin melakukan Scan Handphone Anda ?")
        if scan == QMessageBox.Yes:
            self.circular = SplashScreen()
            self.circular.progress
            self.circular.progressBarValue
            self.result.setText("Root!")
        else:
            print("No")

    def dialog_connect(self):
        connect = QMessageBox.question(
            self, "Connect Button", "Ingin melakukan Connect perangkat ini ?")
        self.result.setText("Not Root!")
        if connect == QMessageBox.Yes:
            selected_item = self.detectDevice.currentText()
            print(selected_item)
            dlg = QMessageBox(self)
            dlg.setText('Berhasil terkoneksi dengan perangkat')
            btn = dlg.exec_()
            if btn == QMessageBox.Ok:
                print('OK')
        else:
            print("No")

    def cek_button(self):
        dlg = QMessageBox()
        dlg.setWindowTitle("Cek Button")
        dlg.setText("Silakan lakukan scan handphone anda terlebih dahulu !")
        dlg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        dlg.setDefaultButton(QMessageBox.No)

        button_clicked = dlg.exec_()
        if button_clicked == QMessageBox.Yes:
            print("Ok")
        else:
            print("No")
            # self.result.setText("Not Root!")

    def exit(self):
        self.app.exit()


# GLOBALS
counter = 0
jumper = 10

# ==> SPLASHSCREEN WINDOW

class SplashScreen(QMainWindow):
    def __init__(self):
        QMainWindow.__init__(self)
        self.ui = Ui_SplashScreen()
        self.ui.setupUi(self)

        # ==> SET INITIAL PROGRESS BAR TO (0) ZERO
        self.progressBarValue(0)

        # ==> REMOVE STANDARD TITLE BAR
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)  # Remove title bar
        # Set background to transparent
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)

        # ==> APPLY DROP SHADOW EFFECT
        self.shadow = QGraphicsDropShadowEffect(self)
        self.shadow.setBlurRadius(20)
        self.shadow.setXOffset(0)
        self.shadow.setYOffset(0)
        self.shadow.setColor(QColor(0, 0, 0, 120))
        self.ui.circularBg.setGraphicsEffect(self.shadow)

        # QTIMER ==> START
        self.timer = QtCore.QTimer()
        self.timer.timeout.connect(self.progress)
        # TIMER IN MILLISECONDS
        self.timer.start(20)

        # SHOW ==> MAIN WINDOW
        ########################
        self.show()
        ## ==> END ##

    # DEF TO LOANDING
    ####################
    def progress(self):
        global counter
        global jumper
        value = counter

        # HTML TEXT PERCENTAGE
        htmlText = """<p><span style=" font-size:68pt;">{VALUE}</span><span style=" font-size:58pt; vertical-align:super;">%</span></p>"""

        # REPLACE VALUE
        newHtml = htmlText.replace("{VALUE}", str(jumper))

        if(value > jumper):
            # APPLY NEW PERCENTAGE TEXT
            self.ui.labelPercentage.setText(newHtml)
            jumper += 10

        # SET VALUE TO PROGRESS BAR
        # fix max value error if > than 100
        if value >= 100:
            value = 1.000
        self.progressBarValue(value)

        # CLOSE SPLASH SCREE AND OPEN APP
        if counter > 101:
            # STOP TIMER
            self.timer.stop()

            # SHOW MAIN WINDOW
            # self.main = MainWindow()
            # self.main.show()

            # CLOSE SPLASH SCREEN
            self.close()

        # INCREASE COUNTER
        counter += 0.5

    # DEF PROGRESS BAR VALUE
    #########################
    def progressBarValue(self, value):

        # PROGRESSBAR STYLESHEET BASE
        styleSheet = """
        QFrame{
        	border-radius: 150px;
        	background-color: qconicalgradient(cx:0.5, cy:0.5, angle:90, stop:{STOP_1} rgba(255, 0, 127, 0), stop:{STOP_2} rgba(85, 170, 255, 255));
        }
        """

        # GET PROGRESS BAR VALUE, CONVERT TO FLOAT AND INVERT VALUES
        # stop works of 1.000 to 0.000
        progress = (100 - value) / 100.0

        # GET NEW VALUES
        stop_1 = str(progress - 0.001)
        stop_2 = str(progress)

        # SET VALUES TO NEW STYLESHEET
        newStylesheet = styleSheet.replace(
            "{STOP_1}", stop_1).replace("{STOP_2}", stop_2)

        # APPLY STYLESHEET WITH NEW VALUES
        self.ui.circularProgress.setStyleSheet(newStylesheet)


# untuk kebutuhan pada saat demo ketika sidang nanti
# if selected_item == self.result.setText("Root"):
#     print("Telah root")

# for i in range(101):
#     self.acquisitionBar.setValue(i)
#     time.sleep(0.03)
# self.convertButton.clicked.connect(self.start_convert_image)

 # workbook = Workbook()
            # sheet = workbook.active

            # max_image_width = sheet.column_dimensions['A'].width
            # max_image_height = sheet.row_dimensions[1].height

            # default_width = 100  # Nilai lebar default
            # default_height = 100  # Nilai tinggi default

            # try:
            #     if max_image_width is not None:
            #         max_image_width = int(max_image_width)
            #     else:
            #         max_image_width = default_width

            #     if max_image_height is not None:
            #         max_image_height = int(max_image_height)
            #     else:
            #         max_image_height = default_height

            #     max_image_size = (max_image_width, max_image_height)
            #     resized_image = image.resize(max_image_size)

            #     image_cell = sheet.cell(row=1, column=1)
            #     image_cell.value = "Image"
            #     sheet.column_dimensions['A'].width = resized_image.width
            #     sheet.row_dimensions[1].height = resized_image.height
            #     sheet.add_image(resized_image, 'A4')

            #     text_cell = sheet.cell(row=2, column=1)
            #     text_cell.value = extracted_text

            #     workbook.save(excel_path)

            #     QMessageBox.information(
            #         self, "Konversi Selesai", "Gambar berhasil dikonversi menjadi file Excel!")
            # except (ValueError, TypeError) as e:
            #     QMessageBox.warning(
            #         self, "Error", "Ukuran lebar atau tinggi gambar tidak valid!\n\n{str(e)}")
